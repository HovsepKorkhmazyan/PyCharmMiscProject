import pandas as pd
import openpyxl


def process_data():
    try:
        with open('db.txt', 'r') as file:
            lines = file.readlines()

        data = []
        for line in lines:
            if line.strip():
                parts = line.strip().split()
                if len(parts) >= 4:
                    data.append({
                        'Name': parts[0],
                        'Surname': parts[1],
                        'Age': int(parts[2]),
                        'Profession': ' '.join(parts[3:])
                    })

        if not data:
            print("No valid data found in db.txt")
            return

        df = pd.DataFrame(data)

        with pd.ExcelWriter('output.xlsx', engine='openpyxl') as writer:
            sorted_df = df.sort_values(by=['Name', 'Surname'])
            sorted_df.to_excel(writer, sheet_name='All Data', index=False)

            workbook = writer.book
            worksheet = writer.sheets['All Data']

            for cell in worksheet['1:1']:
                cell.font = openpyxl.styles.Font(bold=True, color='FFFF00')
                cell.fill = openpyxl.styles.PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

            for row in worksheet.iter_rows(min_row=2, min_col=3, max_col=3):
                for cell in row:
                    if isinstance(cell.value, (int, float)) and cell.value > 25:
                        cell.fill = openpyxl.styles.PatternFill(start_color='00FF00', end_color='00FF00',
                                                                fill_type='solid')

            programmers_df = df[df['Profession'].str.contains('programmer', case=False)]
            programmers_df.to_excel(writer, sheet_name='Programmers', index=False)

            programmers_worksheet = writer.sheets['Programmers']
            for cell in programmers_worksheet['1:1']:
                cell.font = openpyxl.styles.Font(bold=True, color='FFFF00')
                cell.fill = openpyxl.styles.PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

            for row in programmers_worksheet.iter_rows(min_row=2, min_col=3, max_col=3):
                for cell in row:
                    if isinstance(cell.value, (int, float)) and cell.value > 25:
                        cell.fill = openpyxl.styles.PatternFill(start_color='00FF00', end_color='00FF00',
                                                                fill_type='solid')

        print("Excel file 'output.xlsx' created successfully with two sheets.")

    except FileNotFoundError:
        print("Error: db.txt file not found")
    except Exception as e:
        print(f"An error occurred: {str(e)}")


if __name__ == "__main__":
    process_data()
