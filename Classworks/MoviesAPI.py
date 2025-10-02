import datetime
import os
from typing import List

import openpyxl
from fastapi import Depends, FastAPI, HTTPException
from sqlalchemy import Column, Integer, String, create_engine
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.orm import Session, sessionmaker
from pydantic import BaseModel

SQLALCHEMY_DATABASE_URL = "sqlite:///./movies.db"


class MovieBase(BaseModel):
    title: str
    year: int
    genre: str
    director: str


class MovieCreate(MovieBase):
    pass


class Movie(MovieBase):
    id: int

    class Config:
        from_attributes = True


engine = create_engine(
    SQLALCHEMY_DATABASE_URL, connect_args={"check_same_thread": False}
)

SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)

Base = declarative_base()


class DBMovie(Base):
    __tablename__ = "movies"

    id = Column(Integer, primary_key=True, index=True)
    title = Column(String, index=True)
    year = Column(Integer)
    genre = Column(String)
    director = Column(String)


app = FastAPI(
    title="Movie API",
    description="An API to manage movies with XLSX import and export functionality."
)


def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


def populate_db_from_xlsx(db: Session, filepath: str = "films_list_100.xlsx"):
    print(f"Attempting to populate database from '{filepath}'...")
    if not os.path.exists(filepath):
        print(f"'{filepath}' not found. Skipping database population.")
        return

    workbook = openpyxl.load_workbook(filepath)
    sheet = workbook.active

    if db.query(DBMovie).count() > 0:
        print("Database already contains data. Skipping population.")
        return

    movies_added = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue

        title, year_str, genre, director = row

        try:
            year = int(year_str)
        except (ValueError, TypeError):
            print(f"Skipping row due to invalid year: {row}")
            continue

        existing_movie = db.query(DBMovie).filter(DBMovie.title == title).first()
        if not existing_movie:
            db_movie = DBMovie(title=title, year=year, genre=genre, director=director)
            db.add(db_movie)
            movies_added += 1

    db.commit()
    print(f"Successfully added {movies_added} new movies to the database.")


@app.on_event("startup")
def on_startup():
    print("Application starting up...")
    Base.metadata.create_all(bind=engine)
    print("Database tables created (if not exist).")

    db = SessionLocal()
    try:
        populate_db_from_xlsx(db)
    finally:
        db.close()
    print("Startup process finished.")


@app.post("/movies/", response_model=Movie, status_code=201, tags=["Movies"])
def add_movie(movie: MovieCreate, db: Session = Depends(get_db)):
    db_movie = db.query(DBMovie).filter(DBMovie.title == movie.title).first()
    if db_movie:
        raise HTTPException(status_code=400, detail="Movie with this title already exists")

    new_movie = DBMovie(**movie.model_dump())
    db.add(new_movie)
    db.commit()
    db.refresh(new_movie)
    return new_movie


@app.get("/movies/", response_model=List[Movie], tags=["Movies"])
def get_all_movies(skip: int = 0, limit: int = 100, db: Session = Depends(get_db)):
    movies = db.query(DBMovie).offset(skip).limit(limit).all()
    return movies


@app.get("/movies/{movie_id}", response_model=Movie, tags=["Movies"])
def get_movie_by_id(movie_id: int, db: Session = Depends(get_db)):
    movie = db.query(DBMovie).filter(DBMovie.id == movie_id).first()
    if movie is None:
        raise HTTPException(status_code=404, detail="Movie not found")
    return movie


@app.get("/dump-movies-to-xlsx/", status_code=200, tags=["Export"])
def dump_movies_to_xlsx(db: Session = Depends(get_db)):
    movies = db.query(DBMovie).all()
    if not movies:
        raise HTTPException(status_code=404, detail="No movies in the database to dump.")

    current_date = datetime.datetime.now().strftime("%Y-%m-%d")
    filename = f"movies_dump_{current_date}.xlsx"

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Movies Dump"

    header = ["ID", "Title", "Year", "Genre", "Director"]
    sheet.append(header)

    for movie in movies:
        sheet.append([movie.id, movie.title, movie.year, movie.genre, movie.director])

    workbook.save(filename)

    return {"message": f"Successfully dumped {len(movies)} movies to '{filename}'"}
